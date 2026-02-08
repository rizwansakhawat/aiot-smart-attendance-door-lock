"""
Views for Smart Attendance System
=================================
Handles web pages and API endpoints with User Authentication
"""

import json
import base64
import cv2
import numpy as np
import os
import uuid
from datetime import datetime, timedelta
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.utils import timezone

from .models import Student, Attendance, SystemLog, Department
from .services.face_recognition_service import (
    get_face_recognition_service,
)

# Camera Index
CAMERA_INDEX = 0
# Notification Service
try:
    from attendance.services.notification_service import NotificationService
    NOTIFICATIONS_AVAILABLE = True
except:
    NOTIFICATIONS_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════

def is_admin(user):
    """Check if user is admin/staff"""
    return user.is_staff or user.is_superuser


def get_student_for_user(user):
    """Get student profile for logged-in user"""
    try:
        return Student.objects.get(user=user)
    except Student.DoesNotExist:
        return None


# ═══════════════════════════════════════════════════════════════════
# AUTHENTICATION VIEWS
# ═══════════════════════════════════════════════════════════════════

def login_view(request):
    """
    Login page for all users
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        if not username or not password:
            messages.error(request, 'Please enter both username and password')
            return render(request, 'attendance/login.html')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_active:
                login(request, user)
                
                # Log successful login
                SystemLog.objects.create(
                    log_type='success',
                    message=f"User '{username}' logged in successfully"
                )
                
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                
                # Redirect based on user type
                next_url = request.GET.get('next', 'dashboard')
                return redirect(next_url)
            else:
                messages.error(request, 'Your account is disabled. Please contact admin.')
        else:
            # Log failed login
            SystemLog.objects.create(
                log_type='warning',
                message=f"Failed login attempt for username: {username}"
            )
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'attendance/login.html')


def logout_view(request):
    """
    Logout user
    """
    if request.user.is_authenticated:
        username = request.user.username
        logout(request)
        
        SystemLog.objects.create(
            log_type='info',
            message=f"User '{username}' logged out"
        )
        
        messages.success(request, 'You have been logged out successfully.')
    
    return redirect('login')


# ═══════════════════════════════════════════════════════════════════
# USER PROFILE VIEWS
# ═══════════════════════════════════════════════════════════════════

@login_required(login_url='login')
def user_profile(request):
    """
    View user profile - Shows profile information
    """
    user = request.user
    student = get_student_for_user(user)
    
    # Get attendance statistics
    attendance_stats = {}
    if student:
        today = timezone.now().date()
        all_attendance = Attendance.objects.filter(
            student=student,
            entry_type='success'
        )
        
        # Total attendance days
        attendance_stats['total_days'] = all_attendance.values('timestamp__date').distinct().count()
        
        # This month's attendance
        first_of_month = today.replace(day=1)
        attendance_stats['month_attendance'] = all_attendance.filter(
            timestamp__date__gte=first_of_month
        ).values('timestamp__date').distinct().count()
        
        # Last access
        attendance_stats['last_access'] = all_attendance.first()
    
    context = {
        'user': user,
        'student': student,
        'attendance_stats': attendance_stats,
        'is_admin': is_admin(user),
    }
    
    return render(request, 'attendance/profile.html', context)


@login_required(login_url='login')
def update_profile(request):
    """
    Update user profile - Allows users to update their information
    """
    user = request.user
    student = get_student_for_user(user)
    
    if request.method == 'POST':
        # Get form data
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        
        # Password change (optional)
        current_password = request.POST.get('current_password', '')
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')
        
        errors = []
        
        # Validate email uniqueness (excluding current user)
        if email and User.objects.filter(email=email).exclude(pk=user.pk).exists():
            errors.append("This email is already in use by another account.")
        
        # Password validation
        if new_password:
            if not current_password:
                errors.append("Please enter your current password to change it.")
            elif not user.check_password(current_password):
                errors.append("Current password is incorrect.")
            elif new_password != confirm_password:
                errors.append("New passwords do not match.")
            elif len(new_password) < 6:
                errors.append("New password must be at least 6 characters long.")
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return redirect('update_profile')
        
        # Update User model
        user.first_name = first_name
        user.last_name = last_name
        user.email = email
        
        if new_password:
            user.set_password(new_password)
        
        user.save()
        
        # Update Student model if exists
        if student:
            student.email = email if email else student.email
            student.phone = phone if phone else student.phone
            
            # Update name in student if changed
            full_name = f"{first_name} {last_name}".strip()
            if full_name:
                student.name = full_name
            
            # Handle profile photo update
            if 'photo' in request.FILES:
                photo = request.FILES['photo']
                
                # Validate file type
                allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
                if photo.content_type not in allowed_types:
                    messages.warning(request, 'Invalid image format. Please upload JPG, PNG, GIF, or WebP.')
                elif photo.size > 5 * 1024 * 1024:  # 5MB limit
                    messages.warning(request, 'Image too large. Maximum size is 5MB.')
                else:
                    # Delete old photo if exists
                    if student.photo:
                        try:
                            import os
                            if os.path.isfile(student.photo.path):
                                os.remove(student.photo.path)
                        except Exception:
                            pass  # Ignore deletion errors
                    
                    # Save new photo
                    student.photo = photo
            
            # Handle photo removal
            if request.POST.get('remove_photo') == 'true' and student.photo:
                try:
                    import os
                    if os.path.isfile(student.photo.path):
                        os.remove(student.photo.path)
                except Exception:
                    pass
                student.photo = None
            
            student.save()
        
        # Log profile update
        SystemLog.objects.create(
            log_type='info',
            message=f"User '{user.username}' updated their profile"
        )
        
        messages.success(request, 'Your profile has been updated successfully!')
        
        # Re-login if password was changed
        if new_password:
            login(request, user)
            messages.info(request, 'Your password has been changed. You are still logged in.')
        
        return redirect('user_profile')
    
    context = {
        'user': user,
        'student': student,
        'is_admin': is_admin(user),
    }
    
    return render(request, 'attendance/update_profile.html', context)


# ═══════════════════════════════════════════════════════════════════
# DASHBOARD VIEWS
# ═══════════════════════════════════════════════════════════════════

@login_required(login_url='login')
def dashboard(request):
    """
    Dashboard - Shows different view based on user role
    - Admin: Full system dashboard
    - User: Personal attendance dashboard
    """
    if is_admin(request.user):
        return admin_dashboard(request)
    else:
        return user_dashboard(request)


def admin_dashboard(request):
    """
    Admin Dashboard - Full access to all data
    """
    today = timezone.now().date()
    
    # Get today's attendance
    today_attendance = Attendance.objects.filter(
        timestamp__date=today,
        entry_type='success'
    ).select_related('student').order_by('-timestamp')
    
    # Get statistics
    total_students = Student.objects.filter(is_active=True).count()
    present_today = today_attendance.values('student').distinct().count()
    absent_today = total_students - present_today
    
    # Get failed attempts
    failed_attempts = Attendance.objects.filter(
        timestamp__date=today,
        entry_type='denied'
    ).count()
    
    # Get recent entries
    recent_entries = today_attendance[:10]
    
    # Weekly stats for chart
    week_stats = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        count = Attendance.objects.filter(
            timestamp__date=day,
            entry_type='success'
        ).values('student').distinct().count()
        week_stats.append({
            'day': day.strftime('%a'),
            'date': day.strftime('%Y-%m-%d'),
            'count': count
        })
    
    context = {
        'is_admin': True,
        'total_students': total_students,
        'present_today': present_today,
        'absent_today': absent_today,
        'failed_attempts': failed_attempts,
        'recent_entries': recent_entries,
        'today': today,
        'attendance_percentage': round((present_today / total_students * 100), 1) if total_students > 0 else 0,
        'week_stats': week_stats,
    }
    
    return render(request, 'attendance/admin_dashboard.html', context)


def user_dashboard(request):
    """
    User Dashboard - Personal attendance records only
    """
    today = timezone.now().date()
    
    # Get student profile for logged-in user
    student = get_student_for_user(request.user)
    
    if not student:
        messages.warning(request, 'No student profile linked to your account. Please contact admin.')
        context = {
            'is_admin': False,
            'student': None,
            'has_profile': False,
        }
        return render(request, 'attendance/user_dashboard.html', context)
    
    # Get user's attendance records
    all_attendance = Attendance.objects.filter(
        student=student,
        entry_type='success'
    )
    
    # Today's status
    is_present_today = all_attendance.filter(timestamp__date=today).exists()
    
    # Get last access
    last_access = all_attendance.first()  # Already ordered by -timestamp
    
    # This month's attendance
    first_of_month = today.replace(day=1)
    month_attendance = all_attendance.filter(
        timestamp__date__gte=first_of_month
    ).values('timestamp__date').distinct().count()
    
    # Total attendance days
    total_days = all_attendance.values('timestamp__date').distinct().count()
    
    # Recent attendance records
    recent_records = Attendance.objects.filter(student=student).order_by('-timestamp')[:20]
    
    # Weekly attendance for chart
    week_attendance = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        present = all_attendance.filter(timestamp__date=day).exists()
        week_attendance.append({
            'day': day.strftime('%a'),
            'date': day.strftime('%Y-%m-%d'),
            'present': present
        })
    
    context = {
        'is_admin': False,
        'student': student,
        'has_profile': True,
        'is_present_today': is_present_today,
        'last_access': last_access,
        'month_attendance': month_attendance,
        'total_days': total_days,
        'recent_records': recent_records,
        'week_attendance': week_attendance,
        'today': today,
    }
    
    return render(request, 'attendance/user_dashboard.html', context)


# ═══════════════════════════════════════════════════════════════════
# STUDENT REGISTRATION (Admin Only)
# ═══════════════════════════════════════════════════════════════════

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def register_student(request):
    """
    Student registration page - Admin only
    """
    if request.method == 'POST':
        return handle_student_registration(request)
    
    # Fetch active departments from database
    departments = Department.objects.filter(is_active=True).values_list('name', flat=True).order_by('name')
    
    context = {
        'camera_index': CAMERA_INDEX,
        'departments': list(departments),
        'user_types': [
            ('student', 'Student'),
            ('staff', 'Staff'),
            ('visitor', 'Visitor'),
        ]
    }
    
    return render(request, 'attendance/register_student.html', context)

def handle_student_registration(request):
    """
    Handle POST request for student registration
    Also creates a User account and saves profile photo
    """
    try:
        # Get form data
        name = request.POST.get('name', '').strip()
        roll_number = request.POST.get('roll_number', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        department = request.POST.get('department', '').strip()
        user_type = request.POST.get('user_type', 'student')
        face_encodings_json = request.POST.get('face_encodings', '')
        profile_photo_data = request.POST.get('profile_photo', '')  # Base64 image
        create_account = request.POST.get('create_account', 'on')  # Checkbox
        
        # Validation
        errors = []
        
        if not name:
            errors.append("Name is required")
        
        if not roll_number:
            errors.append("Roll Number/ID is required")
        elif Student.objects.filter(roll_number=roll_number).exists():
            errors.append(f"Roll Number '{roll_number}' already exists")
        
        if not face_encodings_json:
            errors.append("Face data is required. Please capture face images.")
        
        if not department:
            errors.append("Department is required")
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return redirect('register_student')
        
        # ═══════════════════════════════════════════════════════════
        # SAVE PROFILE PHOTO FROM BASE64
        # ═══════════════════════════════════════════════════════════
        saved_photo_path = None
        
        if profile_photo_data:
            try:
                # Remove data URL prefix
                if 'base64,' in profile_photo_data:
                    profile_photo_data = profile_photo_data.split('base64,')[1]
                
                # Decode base64
                image_data = base64.b64decode(profile_photo_data)
                
                # Create filename
                filename = f"{roll_number.replace(' ', '_').replace('-', '_')}_{uuid.uuid4().hex[:8]}.jpg"
                
                # Ensure directory exists
                faces_dir = os.path.join(settings.MEDIA_ROOT, 'faces')
                os.makedirs(faces_dir, exist_ok=True)
                
                # Save file
                filepath = os.path.join(faces_dir, filename)
                with open(filepath, 'wb') as f:
                    f.write(image_data)
                
                saved_photo_path = f"faces/{filename}"
                print(f"✅ Profile photo saved: {saved_photo_path}")
                
            except Exception as e:
                print(f"⚠️ Error saving photo: {e}")
                # Continue without photo - not critical
        
        # Get department object
        department_obj = Department.objects.filter(name=department).first()
        
        # Create User account if checkbox is checked
        user = None
        generated_username = None
        generated_password = None

        if create_account:
            # Get custom or generate username
            custom_username = request.POST.get('custom_username', '').strip()
            custom_password = request.POST.get('custom_password', '').strip()
            
            if custom_username:
                generated_username = custom_username.lower().replace(' ', '_')
            else:
                generated_username = roll_number.lower().replace(' ', '').replace('-', '_')
            
            # Check if username exists
            base_username = generated_username
            counter = 1
            while User.objects.filter(username=generated_username).exists():
                generated_username = f"{base_username}_{counter}"
                counter += 1
            
            # Use custom password or generate one
            if custom_password:
                generated_password = custom_password
            else:
                clean_roll = ''.join(c for c in roll_number if c.isalnum())
                generated_password = f"{clean_roll}@123"
            
            # Create user
            user = User.objects.create_user(
                username=generated_username,
                email=email if email else None,
                password=generated_password,
                first_name=name.split()[0] if name else '',
                last_name=' '.join(name.split()[1:]) if len(name.split()) > 1 else ''
            )
            
            print(f"Created user: {generated_username} with password: {generated_password}")  # Debug log
        
        # Create student
        student = Student.objects.create(
            user=user,
            name=name,
            roll_number=roll_number,
            email=email if email else None,
            phone=phone if phone else None,
            department=department_obj if department_obj else None,
            user_type=user_type,
            face_encoding=face_encodings_json,
            photo=saved_photo_path,  # Path to saved image
            is_active=True
        )
        
        # Send welcome email
        if NOTIFICATIONS_AVAILABLE and user and generated_password:
            try:
                NotificationService.notify_registration(student, generated_username, generated_password)
            except Exception as e:
                print(f"Notification error: {e}")
        
        # Refresh face recognition cache
        service = get_face_recognition_service()
        service.refresh_cache()
        
        # Log registration
        SystemLog.objects.create(
            log_type='success',
            message=f"New student registered: {name} ({roll_number})"
        )
        
        # Success message with login credentials
        if user and generated_password:
            messages.success(
                request, 
                f'''
                <div id="success-message" class="p-3">
                    <div class="d-flex align-items-start">
                        <div class="me-3">
                            <span style="font-size: 3rem;">🎉</span>
                        </div>
                        <div class="flex-grow-1">
                            <h4 class="mb-3 text-success">
                                <i class="bi bi-check-circle-fill me-2"></i>Student Registered Successfully!
                            </h4>
                            <p class="mb-3"><strong>{name}</strong> has been added to the system.</p>
                            
                            <div class="card border-primary" style="max-width: 400px;">
                                <div class="card-header bg-primary text-white">
                                    <i class="bi bi-key-fill me-2"></i>Login Credentials
                                </div>
                                <div class="card-body">
                                    <div class="mb-3">
                                        <label class="form-label text-muted small mb-1">Username</label>
                                        <div class="input-group">
                                            <input type="text" class="form-control bg-light" value="{generated_username}" id="copy-username" readonly>
                                            <button class="btn btn-outline-primary" type="button" onclick="copyToClipboard('copy-username')">
                                                <i class="bi bi-clipboard"></i>
                                            </button>
                                        </div>
                                    </div>
                                    <div class="mb-3">
                                        <label class="form-label text-muted small mb-1">Password</label>
                                        <div class="input-group">
                                            <input type="text" class="form-control bg-light" value="{generated_password}" id="copy-password" readonly>
                                            <button class="btn btn-outline-primary" type="button" onclick="copyToClipboard('copy-password')">
                                                <i class="bi bi-clipboard"></i>
                                            </button>
                                        </div>
                                    </div>
                                    <div class="alert alert-warning mb-0 py-2">
                                        <small>
                                            <i class="bi bi-exclamation-triangle-fill me-1"></i>
                                            Please save these credentials! This message will disappear.
                                        </small>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
                <script>
                function copyToClipboard(elementId) {{
                    const input = document.getElementById(elementId);
                    input.select();
                    document.execCommand('copy');
                    
                    const btn = input.nextElementSibling;
                    const originalHtml = btn.innerHTML;
                    btn.innerHTML = '<i class="bi bi-check"></i>';
                    btn.classList.remove('btn-outline-primary');
                    btn.classList.add('btn-success');
                    
                    setTimeout(() => {{
                        btn.innerHTML = originalHtml;
                        btn.classList.remove('btn-success');
                        btn.classList.add('btn-outline-primary');
                    }}, 2000);
                }}
                </script>
                '''
            )
        else:
            messages.success(request, f"✅ Student '{name}' registered successfully!")
        
        return redirect('student_list')
        
    except Exception as e:
        messages.error(request, f"Registration failed: {str(e)}")
        SystemLog.objects.create(
            log_type='error',
            message=f"Registration failed: {str(e)}"
        )
        return redirect('register_student')
# ═══════════════════════════════════════════════════════════════════
# FACE CAPTURE API
# ═══════════════════════════════════════════════════════════════════

@csrf_exempt
@require_http_methods(["POST"])
def capture_face_api(request):
    """
    API endpoint to process captured face image from webcam
    """
    try:
        data = json.loads(request.body)
        image_data = data.get('image', '')
        
        if not image_data:
            return JsonResponse({
                'success': False,
                'error': 'No image data received'
            })
        
        # Remove data URL prefix if present
        if 'base64,' in image_data:
            image_data = image_data.split('base64,')[1]
        
        # Decode base64 image
        image_bytes = base64.b64decode(image_data)
        nparr = np.frombuffer(image_bytes, np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if image is None:
            return JsonResponse({
                'success': False,
                'error': 'Could not decode image'
            })
        
        # Get face recognition service
        service = get_face_recognition_service()
        
        # Detect face
        face_location = service.detect_single_face(image)
        
        print(f"Detected face location: {face_location}")  # Debug log
        
        if face_location is None:
            return JsonResponse({
                'success': False,
                'error': 'No face detected. Please position your face in the frame.'
            })
        
        # Check if face is valid
        if not service.is_face_valid(face_location):
            return JsonResponse({
                'success': False,
                'error': 'Face is too small. Please move closer to the camera.'
            })
        
        # Generate encoding
        encoding = service.generate_encoding(image, face_location)
        
        if encoding is None:
            return JsonResponse({
                'success': False,
                'error': 'Could not generate face encoding'
            })
        
        return JsonResponse({
            'success': True,
            'encoding': encoding.tolist(),
            'face_location': {
                'top': face_location[0],
                'right': face_location[1],
                'bottom': face_location[2],
                'left': face_location[3]
            },
            'message': 'Face captured successfully!'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@csrf_exempt
@require_http_methods(["GET"])
def check_camera_api(request):
    """
    API endpoint to check if camera is available
    """
    try:
        cap = cv2.VideoCapture(CAMERA_INDEX)
        
        if cap.isOpened():
            ret, frame = cap.read()
            cap.release()
            
            if ret:
                return JsonResponse({
                    'success': True,
                    'camera_index': CAMERA_INDEX,
                    'message': 'Camera is available'
                })
        
        return JsonResponse({
            'success': False,
            'error': 'Camera not available.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


# ═══════════════════════════════════════════════════════════════════
# STUDENT MANAGEMENT (Admin Only)
# ═════════════════════════════════════════════════��═════════════════

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def student_list(request):
    """
    List all registered students - Admin only
    """
    search = request.GET.get('search', '')
    department = request.GET.get('department', '')
    user_type = request.GET.get('user_type', '')
    status = request.GET.get('status', '')
    
    students = Student.objects.all().order_by('-registered_at')
    
    if search:
        students = students.filter(
            Q(name__icontains=search) |
            Q(roll_number__icontains=search) |
            Q(email__icontains=search)
        )
    
    if department:
        students = students.filter(department__name__iexact=department)
    
    if user_type:
        students = students.filter(user_type__iexact=user_type)
    
    if status:
        if status == 'active':
            students = students.filter(is_active=True)
        elif status == 'inactive':
            students = students.filter(is_active=False)
    
    paginator = Paginator(students, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get unique active departments from database
    departments = Department.objects.filter(is_active=True).values_list('name', flat=True).order_by('name')
    
    context = {
        'page_obj': page_obj,
        'students': page_obj,
        'total_students': students.count(),
        'departments': list(departments),
        'search': search,
        'selected_department': department,
        'selected_user_type': user_type,
        'selected_status': status,
    }
    
    return render(request, 'attendance/student_list.html', context)


@login_required(login_url='login')
def student_detail(request, pk):
    """
    View student details - Admin sees any student, User sees only self
    """
    student = get_object_or_404(Student, pk=pk)
    
    # Check permission: Admin can see all, User can see only their own
    if not is_admin(request.user):
        user_student = get_student_for_user(request.user)
        if not user_student or user_student.pk != student.pk:
            messages.error(request, 'You do not have permission to view this profile.')
            return redirect('dashboard')
    
    all_attendance = Attendance.objects.filter(
        student=student,
        entry_type='success'
    )
    
    total_days = all_attendance.values('timestamp__date').distinct().count()
    
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_attendance = Attendance.objects.filter(
        student=student,
        timestamp__gte=thirty_days_ago,
        entry_type='success'
    ).values('timestamp__date').distinct().count()
    
    attendance_records = Attendance.objects.filter(
        student=student
    ).order_by('-timestamp')[:50]
    
    context = {
        'student': student,
        'attendance_records': attendance_records,
        'total_attendance_days': total_days,
        'recent_attendance': recent_attendance,
        'is_admin': is_admin(request.user),
    }
    
    return render(request, 'attendance/student_detail.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def delete_student(request, pk):
    """
    Delete a student - Admin only
    """
    try:
        student = Student.objects.get(pk=pk)
    except Student.DoesNotExist:
        messages.error(request, f"Student with ID {pk} does not exist or has already been deleted.")
        return redirect('student_list')
    
    if request.method == 'POST':
        name = student.name
        
        # Also delete linked user account
        if student.user:
            student.user.delete()
        
        student.delete()
        
        service = get_face_recognition_service()
        service.refresh_cache()
        
        SystemLog.objects.create(
            log_type='warning',
            message=f"Student deleted: {name}"
        )
        
        messages.success(request, f"Student '{name}' has been deleted.")
        return redirect('student_list')
    
    return render(request, 'attendance/delete_student.html', {'student': student})


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def toggle_student_status(request, pk):
    """
    Toggle student active/inactive status - Admin only
    """
    student = get_object_or_404(Student, pk=pk)
    student.is_active = not student.is_active
    student.save()
    
    # Also toggle user account if exists
    if student.user:
        student.user.is_active = student.is_active
        student.user.save()
    
    service = get_face_recognition_service()
    service.refresh_cache()
    
    status = "activated" if student.is_active else "deactivated"
    messages.success(request, f"Student '{student.name}' has been {status}.")
    
    return redirect('student_list')


# ═══════════════════════════════════════════════════════════════════
# ATTENDANCE RECORDS
# ═══════════════════════════════════════════════════════════════════

@login_required(login_url='login')
def attendance_list(request):
    """
    View attendance records
    - Admin: All records
    - User: Only their own records
    """
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    student_id = request.GET.get('student', '')
    entry_type = request.GET.get('entry_type', '')
    
    # Base queryset based on user role
    if is_admin(request.user):
        records = Attendance.objects.all().select_related('student').order_by('-timestamp')
    else:
        student = get_student_for_user(request.user)
        if student:
            records = Attendance.objects.filter(student=student).order_by('-timestamp')
        else:
            records = Attendance.objects.none()
    
    # Apply filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            records = records.filter(timestamp__date__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            records = records.filter(timestamp__date__lte=date_to_obj)
        except:
            pass
    
    if student_id and is_admin(request.user):
        records = records.filter(student_id=student_id)
    
    if entry_type:
        records = records.filter(entry_type=entry_type)
    
    paginator = Paginator(records, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    students = Student.objects.filter(is_active=True).order_by('name') if is_admin(request.user) else []
    
    context = {
        'page_obj': page_obj,
        'records': page_obj,
        'total_records': records.count(),
        'students': students,
        'date_from': date_from,
        'date_to': date_to,
        'selected_student': student_id,
        'selected_entry_type': entry_type,
        'is_admin': is_admin(request.user),
    }
    
    return render(request, 'attendance/attendance_list.html', context)


# ═══════════════════════════════════════════════════════════════════
# REPORTS (Admin Only)
# ═══════════════════════════════════════════════════════════════════

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def reports(request):
    """
    Reports page - Admin only
    """
    context = {
        'students': Student.objects.filter(is_active=True).order_by('name'),
    }
    return render(request, 'attendance/reports.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def generate_report(request):
    """
    Generate attendance report - Admin only
    """
    if request.method != 'POST':
        return redirect('reports')
    
    report_type = request.POST.get('report_type', 'daily')
    date_from = request.POST.get('date_from', '')
    date_to = request.POST.get('date_to', '')
    student_id = request.POST.get('student', '')
    format_type = request.POST.get('format', 'html')
    
    records = Attendance.objects.filter(entry_type='success').select_related('student')
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            records = records.filter(timestamp__date__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            records = records.filter(timestamp__date__lte=date_to_obj)
        except:
            pass
    
    if student_id:
        records = records.filter(student_id=student_id)
    
    records = records.order_by('-timestamp')
    
    summary = records.values('student__name', 'student__roll_number').annotate(
        total_days=Count('timestamp__date', distinct=True)
    ).order_by('student__name')
    
    context = {
        'records': records[:500],
        'summary': summary,
        'date_from': date_from,
        'date_to': date_to,
        'total_records': records.count(),
    }
    
    if format_type == 'excel':
        return generate_excel_report(records, summary, request)
    
    return render(request, 'attendance/report_result.html', context)


def generate_excel_report(records, summary, request):
    """
    Generate Excel report
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        headers = ['Sr.', 'Name', 'Roll Number', 'Date', 'Time', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row, record in enumerate(records[:1000], 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=record.student.name if record.student else 'Unknown')
            ws.cell(row=row, column=3, value=record.student.roll_number if record.student else 'N/A')
            ws.cell(row=row, column=4, value=record.timestamp.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=5, value=record.timestamp.strftime('%H:%M:%S'))
            ws.cell(row=row, column=6, value='Present')
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f"Error generating Excel report: {str(e)}")
        return redirect('reports')


# ═══════════════════════════════════════════════════════════════════
# SYSTEM LOGS (Admin Only)
# ═══════════════════════════════════════════════════════════════════

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def system_logs(request):
    """
    View system logs - Admin only
    """
    log_type = request.GET.get('type', '')
    
    logs = SystemLog.objects.all().order_by('-timestamp')
    
    if log_type:
        logs = logs.filter(log_type=log_type)
    
    paginator = Paginator(logs, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'logs': page_obj,
        'selected_type': log_type,
    }
    
    return render(request, 'attendance/system_logs.html', context)


# ═══════════════════════════════════════════════════════════════════
# API ENDPOINTS
# ═══════════════════════════════════════════════════════════════════

@csrf_exempt
def api_dashboard_stats(request):
    """
    API endpoint to get dashboard statistics
    """
    today = timezone.now().date()
    
    total_students = Student.objects.filter(is_active=True).count()
    present_today = Attendance.objects.filter(
        timestamp__date=today,
        entry_type='success'
    ).values('student').distinct().count()
    
    failed_attempts = Attendance.objects.filter(
        timestamp__date=today,
        entry_type='denied'
    ).count()
    
    recent = Attendance.objects.filter(
        timestamp__date=today,
        entry_type='success'
    ).select_related('student').order_by('-timestamp')[:5]
    
    recent_list = []
    for entry in recent:
        recent_list.append({
            'name': entry.student.name if entry.student else 'Unknown',
            'time': entry.timestamp.strftime('%H:%M:%S'),
            'roll_number': entry.student.roll_number if entry.student else 'N/A'
        })
    
    return JsonResponse({
        'total_students': total_students,
        'present_today': present_today,
        'absent_today': total_students - present_today,
        'failed_attempts': failed_attempts,
        'recent_entries': recent_list,
    })